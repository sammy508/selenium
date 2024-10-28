from selenium import webdriver
import openpyxl
import pandas as pnd
import glob as glob
filepath = r"D:\selenium projects\selenium\1st-proj OrangeHRM\Data driven testing\exceldata.xlsx"  # r"vbbb" denotes raw string
workbook = openpyxl.load_workbook(filepath)
sheet = workbook['Sheet 1']  # for specific sheet
sheet = workbook.active  # can use this if we have single active sheet

rows = sheet.max_row
colms = sheet.max_column

for row in range(1, rows+1):
    for col in range(1, colms+1):
        print(sheet.cell(row,col).value, end="    ")
    print()    





