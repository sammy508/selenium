import openpyxl
from openpyxl.styles import PatternFill

def row_count(filename, SheetName):
    workbook = openpyxl.load_workbook(filename)
    # sheet  = workbook.active
    sheet = workbook[SheetName]
    return (sheet.max_row)

def row_count(filename, SheetName):
    workbook = openpyxl.load_workbook(filename)
    # sheet  = workbook.active
    sheet = workbook[SheetName]
    return (sheet.max_column)


def read_data(filename,SheetName, rowNo, colNo):
    workbook = openpyxl.load_workbook(filename)
    # sheet  = workbook.active
    sheet = workbook[SheetName]
    return sheet.cell(rowNo,colNo).value


def write_data(filename,SheetName, rowNo, colNo, data):
    workbook = openpyxl.load_workbook(filename)
    # sheet  = workbook.active
    sheet = workbook[SheetName]
    sheet.cell(rowNo,colNo).value = data
    workbook.save(filename)

def fillgreen_color(filename,SheetName, rowNo, colNo)  :
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[SheetName]
    greenfill = PatternFill(start_color="60b212",
                            end_color= "60b212",
                            fill_type= "solid"
                          
                            )
    workbook.save(filename)
    

def fillred_color(filename,SheetName, rowNo, colNo)  :
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[SheetName]
    greenfill = PatternFill(start_color="ff0000",
                            end_color= "ff0000",
                            fill_type= "solid"
                          
                            )
    workbook.save(filename)



