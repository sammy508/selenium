from selenium import webdriver
import openpyxl
import pandas as pnd
import glob as glob
filepath = r"D:\selenium projects\selenium\1st-proj OrangeHRM\Data driven testing\exceldata.xlsx"  # r"vbbb" denotes raw string
workbook = openpyxl.load_workbook(filepath)
sheet = workbook['Sheet 1']  # for specific sheet
sheet = workbook.active  # can use this if we have single active sheet

rows = sheet.max_row
colms = sheet.max_column

for row in range(1, rows+1):
    for col in range(1, colms+1):
        print(sheet.cell(row,col).value, end="    ")
    print()    





# # Specify the path to your Excel file
# filepath = r"D:\selenium projects\selenium\1st-proj OrangeHRM\Data driven testing\exceldata.xlsx"

# # Use pandas to read the Excel file
# try:
#     df = pd.read_excel(filepath, sheet_name="Sheet1")  # Adjust the sheet name if necessary
#     print("Number of rows in the Excel file:", len(df))
    
#     # Print the DataFrame contents
#     print("\nContents of the Excel file:")
#     print(df)

# except FileNotFoundError:
#     print("Error: The file was not found. Please check the file path.")
# except ValueError as ve:
#     print(f"Value Error: {ve}")
# except Exception as e:
#     print(f"An error occurred: {e}")

# # Optionally, you can use openpyxl to get more detailed access to the sheet
# try:
#     workbook = openpyxl.load_workbook(filepath)
#     sheet = workbook['Sheet 1']  # Adjust the sheet name if necessary

#     print("\nDetailed cell values from openpyxl:")
#     for row in sheet.iter_rows(values_only=True):
#         print(row)

# except Exception as e:
#     print(f"An error occurred while accessing the workbook: {e}")
